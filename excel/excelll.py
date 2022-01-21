"""
File: excelll.py
Author: Jin Lexuan
E-mail: jlx321@126.com
Time: 2020-05-24 14:29:42
Function:


"""
import openpyxl
import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 520
ws.append([1, 2, 3])
ws['A3'] = datetime.datetime.now()
wb.save('demo.xlsx')
